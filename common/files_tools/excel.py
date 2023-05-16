# -*- coding: utf-8 -*-
# !/usr/bin/python
# @Date : 2023-2-19
# @Author : chenyongzhi
# @File : do_excel.py
# @Software: VS
# @Desc: Excel操作类
import json
import os
import sys

# 把当前目录加入到系统环境变量中
sys.path.append(os.path.split(os.path.dirname(os.path.abspath(__file__)))[0])
sys.path.append("../..")
sys.path.append('venv/Lib/site-packages')
from openpyxl import load_workbook, Workbook


class DoExcel:
    """
    excel操作类
    """

    def __init__(self, path=None, index_table=0):
        """
        :param: path(str):文件路径，
        1. 为None时：调用save()方法时需要传入filename
        2. 如果不为空表示打开已有文件
        :param: index_table(int):excel中的那个单元
        """
        if path:
            self.wb = load_workbook(path)
        else:
            self.wb = Workbook()
        self.path = path
        sheets = self.wb.sheetnames
        sheet = sheets[index_table]
        self.sheet = self.wb[sheet]
        self.cell = self.sheet.cell

    def set_value_by_cell(self, row, column, value):
        """
        通过cell设置值
        :param: row(int):行
        :param: column(int):列
        :param: value(str):设置值
       """
        self.cell(row, column).value = value

    def set_value_by_table(self, tag, value):
        """
        通过A1坐标设置值
        例如：设置A1的值为hello
        set_value_by_table('A1', 'hello')
        :param: tag(str):具体坐标
        :param: value(str):值
        """
        self.sheet[tag] = value

    def get_value_by_table(self, tag):
        """
        通过A1坐标获取值
        例如：获取A1的值为hello
        get_value_by_table('A1')
        :param: tag(str):具体坐标
        :param: value(str):值
        :return <Cell 'Sheet'.A1>对象
        """
        return self.sheet[tag].value

    def get_value_by_cell(self, row, column):
        """
        通过cell坐标设置值 row行 column 列
        例如：获取第一行第一列的值
        get_value_by_cell(self,1,1)
        :param: row(str):行
        :param: column(str):列
        """
        return self.cell(row, column).value

    def get_max_row(self):
        """获取最大行数 """
        return self.sheet.max_row

    def get_max_col(self):
        """获取最大列数 """
        return self.sheet.max_column

    def get_col_value(self, column, row_start=1, row_end=None):
        """
        获取某列多少行的值，默认为所有
        introduce
        :param: column(int):第几列
        :param: row_start(int):开始的行，默认第一行
        :param: row_end(int):结束的行，默认获取全部
        :return list()
        """
        if not row_end:
            row_end = self.get_max_row()
        column_data = []
        for i in range(row_start, row_end + 1):
            cell_value = self.cell(row=i, column=column).value
            column_data.append(cell_value)
        return column_data

    # 获取某行所有值
    def get_row_value(self, row, col_start=1, col_end=None):
        """
        获取某行多少列的值，默认为所有
        introduce
        :param: row(int):第几行
        :param: col_start(int):开始的列，默认第一列
        :param: col_end(int):结束的列，默认获取全部
        :return list()
        """
        if not col_end:
            col_end = self.get_max_col()
        row_data = []
        for i in range(col_start, col_end + 1):
            cell_value = self.cell(row=row, column=i).value
            row_data.append(cell_value)
        return row_data

    def create_sheet(self, title):
        return self.wb.create_sheet(title=title)

    def copy_sheet(self, source_sheet, destination_sheet):
        source = self.wb[source_sheet]
        destination = self.wb[destination_sheet]
        for row in source.iter_rows(values_only=True):
            destination.append(row)

    def save(self, filename=None):
        """
        获取文件名
        :param: filename(str):保存的文件名
        :return:bool
        """
        if filename is None and self.path is None:
            print("保存失败：没有设置文件名")
            return False
        self.wb.save((filename if filename.endswith(".xlsx") else filename + ".xlsx") if filename else self.path)
        print("保存成功")
        return True

    def other_sheet_save(self, filename=None):
        if filename is None and self.path is None:
            print("保存失败：没有设置文件名")
            return False
        # 如果当前sheet有值，则另开一个sheet保存数据
        if self.get_max_row():
            pass

    def do_main(self, filename, *d, output_filename=None, copy_sheets=True):
        """
        动态保存列表嵌套字典的数据到 excel 中
        :param:*d(list):传入的数据列表
        :param copy_sheets: 是否复制源文件的其他 sheet，默认为 True
        :param *d: 传入的数据列表
        :return:bool
        """
        # 复制源文件的其他 sheet 到新的 Excel 文件
        if copy_sheets:
            for sheet_name in self.wb.sheetnames:
                if sheet_name != self.wb.active.title:
                    self.create_sheet(sheet_name)
                    self.copy_sheet(sheet_name, sheet_name)

        for i, val in enumerate(d):
            c = 0
            for k, v in val.items():
                # key 作为第一行标题写入
                self.set_value_by_cell(1, c + 1, k)
                # value 作为每一条数据写入
                self.set_value_by_cell(i + 2, c + 1, v)
                c += 1
        if output_filename:
            self.save(output_filename)
            return
        self.save(filename)


if __name__ == '__main__':
    ex = DoExcel()
    BASE_PATH = os.path.dirname(os.path.abspath(__file__))  # 获取当前文件所在的文件夹路径
    data = [{"url": "1234", "header": "2134", "method": "get", "body": "hhh", "ok": 12345},
            {"url": "1234", "header": "2134"},
            {"url": "1234", "header": "2134", "method": "{}sss", "body": json.dumps({})}]
    ex.do_main("excel.xlsx", *data)
