# -*- coding: utf-8 -*-
# !/usr/bin/python
# @Date : 2023-2-19
# @Author : chenyongzhi
# @File : excel.py
# @Desc: Excel操作类
import json
import os
import sys

# 把当前目录加入到系统环境变量中
sys.path.append(os.path.split(os.path.dirname(os.path.abspath(__file__)))[0])
sys.path.append("../..")
# sys.path.append('venv/Lib/site-packages')
from openpyxl import load_workbook, Workbook
from common.file_handling import logger


class DoExcel:
    """
    excel操作类
    """

    def __init__(self, path=None, index_table=0):
        """
        :param: path(str):需要打开的文件（模板文件/源文件），
        1. 为None时：调用save()方法时需要传入filename
        2. 如果不为空表示打开已有文件
        :param: index_table(int):excel中的那个单元
        """
        if path:
            self.wb = load_workbook(path)
        else:
            self.wb = Workbook()
        self.path = path
        self.index_table = index_table
        # sheets = self.wb.sheetnames
        # sheet = sheets[self.index_table]
        # self.sheet = self.wb[sheet]
        # self.cell = self.sheet.cell

    def set_value_by_cell(self, sheet, row, column, value):
        """
        通过cell设置值
        :param: row(int):行
        :param: column(int):列
        :param: value(str):设置值
       """
        ws = self.wb[sheet]
        ws.cell(row, column).value = value

    def set_value_by_table(self, sheet, tag, value):
        """
        通过A1坐标设置值
        例如：设置A1的值为hello
        set_value_by_table('A1', 'hello')
        :param: tag(str):具体坐标
        :param: value(str):值
        """
        ws = self.wb[sheet]
        ws.sheet[tag] = value

    def get_value_by_table(self, sheet, tag):
        """
        通过A1坐标获取值
        例如：获取A1的值为hello
        get_value_by_table('A1')
        :param: tag(str):具体坐标
        :param: value(str):值
        :return <Cell 'Sheet'.A1>对象
        """
        ws = self.wb[sheet]
        return ws.sheet[tag].value

    def get_value_by_cell(self, sheet, row, column):
        """
        通过cell坐标设置值 row行 column 列
        例如：获取第一行第一列的值
        get_value_by_cell(self,1,1)
        :param: row(str):行
        :param: column(str):列
        """
        ws = self.wb[sheet]
        return ws.cell(row, column).value

    def get_max_row(self, sheet):
        """获取最大行数 """
        ws = self.wb[sheet]
        return ws.sheet.max_row

    def get_max_col(self, sheet):
        """获取最大列数 """
        ws = self.wb[sheet]
        return ws.sheet.max_column

    def get_col_value(self, sheet, column, row_start=1, row_end=None):
        """
        获取某列多少行的值，默认为所有
        introduce
        :param: column(int):第几列
        :param: row_start(int):开始的行，默认第一行
        :param: row_end(int):结束的行，默认获取全部
        :return list()
        """
        ws = self.wb[sheet]
        if not row_end:
            row_end = self.get_max_row(sheet)
        column_data = []
        for i in range(row_start, row_end + 1):
            cell_value = ws.cell(row=i, column=column).value
            column_data.append(cell_value)
        return column_data

    # 获取某行所有值
    def get_row_value(self, sheet, row, col_start=1, col_end=None):
        """
        获取某行多少列的值，默认为所有
        introduce
        :param: row(int):第几行
        :param: col_start(int):开始的列，默认第一列
        :param: col_end(int):结束的列，默认获取全部
        :return list()
        """
        ws = self.wb[sheet]
        if not col_end:
            col_end = self.get_max_col(sheet)
        row_data = []
        for i in range(col_start, col_end + 1):
            cell_value = ws.cell(row=row, column=i).value
            row_data.append(cell_value)
        return row_data

    @logger.log_decorator()
    def save(self, filename=None):
        """
        获取文件名
        :param: filename(str):保存的文件名
        :return:bool
        """
        if filename:
            self.wb.save(filename)
        elif self.path:
            self.wb.save(self.path)
        else:
            # print("保存失败：没有设置文件名")
            return False
        # print("保存成功")
        return True

    @logger.log_decorator()
    def do_main(self, output_filename=None, *data):
        """
        动态保存列表嵌套字典的数据到 excel 中
        :param: output_filename(str):另存为的文件名
        :param: *data(list):传入的数据列表
        :return: bool
        """
        sheet_names = self.wb.sheetnames
        target_sheet = sheet_names[self.index_table]
        ws = self.wb[target_sheet]

        # 清空目标sheet除第一行外的数据
        ws.delete_rows(2, ws.max_row)

        for i, val in enumerate(data):
            c = 0
            for k, v in val.items():
                # key 作为第一行标题写入
                self.set_value_by_cell(target_sheet, 1, c + 1, k)
                # value 作为每一条数据写入
                self.set_value_by_cell(target_sheet, i + 2, c + 1, v)
                c += 1

        self.save(output_filename)


if __name__ == '__main__':
    data = [{"url": "1234", "header": "2134", "method": "get", "body": "hhh", "ok": 12345},
            {"url": "1234", "header": "2134"},
            {"url": "1234", "header": "2134", "method": "{}sss", "body": json.dumps({})}]
    from common.config import Config

    template_file = Config.templates
    excel = DoExcel(path=template_file)
    out_file = r'..\..\OutPut\另存为的文件.xlsx'
    excel.do_main(out_file, *data)
