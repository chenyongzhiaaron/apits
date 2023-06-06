#!/usr/bin/env python
# encoding: utf-8
"""
@author: kira
@contact: 262667641@qq.com
@file: excel_handler.py
@time: 2023/6/5 11:09
@desc:
"""
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


class ExcelHandler():
    '''操作Excel'''

    def __init__(self, file):
        '''初始化函数'''
        self.file = file

    def open_sheet(self, sheet_name) -> Worksheet:
        '''打开表单'''
        wb = load_workbook(self.file)
        sheet = wb[sheet_name]
        return sheet

    def read_header(self, sheet_name):
        '''获取表单的表头'''
        sheet = self.open_sheet(sheet_name)
        headers = []
        for i in sheet[1]:
            headers.append(i.value)
        return headers

    def read_rows(self, sheet_name):
        '''读取除表头外所有数据（除第一行外的所有数据）'''
        sheet = self.open_sheet(sheet_name)
        rows = list(sheet.rows)[1:]
        data = []
        for row in rows:
            row_data = []
            for cell in row:
                row_data.append(cell.value)
            data.append(row_data)
        return data

    def read_key_value(self, sheet_name):
        '''获取所有数据，且将表头中的内容与数据结合展示（以字典的形式）'''
        sheet = self.open_sheet(sheet_name)
        rows = list(sheet.rows)
        # 获取标题
        data = []
        for row in rows[1:]:
            rwo_data = []
            for cell in row:
                rwo_data.append(cell.value)
            # 列表转换成字典，与表头里的内容使用zip函数进行打包
            data_dict = dict(zip(self.read_header(sheet_name), rwo_data))
            data.append(data_dict)
        return data

    @staticmethod
    def write_change(file, sheet_name, row, column, data):
        '''写入Excel数据'''
        wb = load_workbook(file)
        sheet = wb[sheet_name]
        # 修改单元格
        sheet.cell(row, column).value = data
        # 保存
        wb.save(file)
        # 关闭
        wb.close()
