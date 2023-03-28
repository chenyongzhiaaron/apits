# -*- coding: utf-8 -*-
# @Time : 2021/5/17 19:56
# @Author : kira
# @Email : 262667641@qq.com
# @File : JSO.py
# @Project : api-test-project

import json
import random
import time

from openpyxl import load_workbook


#  当前有文件case.xlsx,设计程序将excel中的用例读取到一个生成器

def gen_red(file_path, sheet_name):
    """
    读取excel数据的方法
    Args:
        file_path:文件路劲
        sheet_name:表名

    Returns:读取到的数据

    """
    workbook = load_workbook(file_path)
    sheet = workbook[sheet_name]
    row_max = sheet.max_row  # 最大行数
    column_max = sheet.max_column  # 最多列数
    for row in range(2, row_max + 1):
        row_dict = {} # 每行数据
        for col in range(1,column_max+1):
            # key 等于表头 value 具体值
            row_dict[sheet.cell(row=1,column=col).value] = sheet.cell(row=row)
        yield row_dict
